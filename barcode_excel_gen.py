import os
import barcode
from barcode.writer import ImageWriter
import openpyxl
from openpyxl.drawing.image import Image as opImage
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder, RowDimension
import clothing


def column_parse(ws):  # Find locs of desired columns
    type_obj = size_code = art_pos = art_wb = art_imt = color = size = bar = price = compl = 1000
    for x in range(1, ws.max_column):
        name = ws.cell(1, x).value
        if name == 'Предмет':
            type_obj = x
        elif name == 'Код размера (chrt_id)':
            size_code = x
        elif name == 'Артикул поставщика':
            art_pos = x
        elif name == 'Артикул WB':
            art_wb = x
        elif name == 'Артикул ИМТ':
            art_imt = x
        elif name == 'Артикул Цвета':
            color = x
        elif name == 'Размер':
            size = x
        elif name == 'Баркод':
            bar = x
        elif name == 'Розничная цена, руб':
            price = x
        elif name == 'Комплектация':
            compl = x
    return type_obj, size_code, art_pos, art_wb, art_imt, color, size, bar, price, compl


def parse(file):
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook.active
    art_pos = []
    size = []
    bar = []
    types = []
    color = []
    art_WB = []
    col_num = column_parse(worksheet)
    for row in range(2, int(worksheet.max_row)+1, 1):

        # Артикул изделия, без цвета
        art = worksheet.cell(row=row, column=col_num[4]).value
        if art is None:
            art = '-'

        # Цветовой код
        color_code = worksheet.cell(row=row, column=col_num[5]).value
        if color_code is None:
            color_code = '-'
        color.append(clothing.colors(color_code, art))

        # Артикул поставщика
        if worksheet.cell(row=row, column=col_num[2]).value is None:
            art_pos.append('-')
        else:
            art_pos.append(worksheet.cell(row=row, column=col_num[2]).value)

        # Размер
        if worksheet.cell(row=row, column=col_num[6]).value is None:
            size.append('-')
        else:
            size.append(worksheet.cell(row=row, column=col_num[6]).value)

        # Баркод
        if worksheet.cell(row=row, column=col_num[7]).value is None:
            pass
        else:
            bar.append(worksheet.cell(row=row, column=col_num[7]).value)

        # Тип изделия
        if worksheet.cell(row=row, column=col_num[0]).value is None:
            types.append('-')
        else:
            types.append(clothing.rewrite(worksheet.cell(row=row, column=col_num[0]).value))

        # Артикул WB
        if worksheet.cell(row=row, column=col_num[3]).value is None:
            art_WB.append('-')
        else:
            art_WB.append(worksheet.cell(row=row, column=col_num[3]).value)

    return art_pos, size, bar, types, color, art_WB


def pathSplit(file): # Split path to the file into the parts
    directory, path = os.path.split(file)
    filename, ext = os.path.splitext(path)
    return directory, path, filename, ext


def row_num(ws):
    return int(ws.max_row)


def folderCreate(directory):  # Creating folder with barcodes
    try:
        os.makedirs(directory + '/folder')
    except IOError:
        pass


def gen_excel(full_path, code):
    directory, path, filename, ext = pathSplit(full_path)
    wb = openpyxl.load_workbook(full_path)
    ws = wb.active
    # Changing the dimensions of the cells
    col_holder = DimensionHolder(worksheet=ws)
    row_holder = DimensionHolder(worksheet=ws)

    col = ws.max_column + 1


    # Creating barcodes and adding to the worksheet
    folderCreate(directory)

    columns = column_parse(ws)

    for k in range(2, row_num(ws)+1, 1):
        os.chdir(directory)

        cell = ws.cell(row=k, column=columns[7])
        img = opImage(barcode.get(code, str(cell.value), writer=ImageWriter()).save(
                'folder/' + code + '_' + str(cell.value), {
                "module_width":0.35, "module_height": 10, "font_size": 10, "text_distance": 1, "quiet_zone": 3
            }))
        ws.add_image(img, str(get_column_letter(col)) + str(k))

        # Changing sizes of the cells

        row_holder[k] = RowDimension(ws, index=1, height=img.height)

    # Writing new dimensions
    print(img.height, img.width)
    ws.row_dimensions = row_holder
    col_holder[get_column_letter(col)] = ColumnDimension(ws, index=str(get_column_letter(col)), width=img.width*0.1)
    ws.column_dimensions = col_holder
    # Creating new excel file with barcodes
    os.chdir(directory)
    wb.save(filename+'_'+code+ext)
